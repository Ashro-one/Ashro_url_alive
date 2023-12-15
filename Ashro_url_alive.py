import openpyxl
from openpyxl import Workbook
import argparse
import threading
import requests
import warnings
import os
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from urllib3.exceptions import InsecureRequestWarning

#Data:2023/12/15
#Author:Ashro

#去除ssl警告
warnings.filterwarnings(action='ignore', category=InsecureRequestWarning, module='urllib3')

#避免多线程之间的条件竞争
class ThreadLocalData:
    def __init__(self):
        self.visited_urls = set()

# 使用线程本地变量
thread_local_data = threading.local()
thread_local_data.data = ThreadLocalData()

# 用于确保线程安全的锁
lock = threading.Lock()

#进行网站存活检测
def get_response_header(url):
    try:
        # 设置请求头
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "close"
        }

        # 发送GET请求获取页面内容
        response = requests.get(url, timeout=5, headers=headers, verify=False)

        # 使用BeautifulSoup解析页面内容
        soup = BeautifulSoup(response.text, 'html.parser')

        # 提取页面标题
        title = soup.title.string.strip() if soup.title else None

        # 获取响应状态码、服务器信息和重定向地址
        code = response.status_code     #获取网站响应状态码
        server = response.headers.get('Server', None)       #获取网站服务器信息
        location = response.headers.get('Location', None)       #获取网站是否存在重定向

        # 将获取到的信息组装成列表并返回
        headers = [code, title, server, location]
        return headers

    except Exception as e:
        # 如果发生异常，返回None
        return None


def process_url(url, output_file):
    """
    处理给定的URL，获取响应头信息，将相关信息输出到控制台，并将信息保存到Excel文件。
    参数:
    - url (str): 要处理的URL。
    - output_file (str): 保存输出信息的Excel文件路径。
    """
    # 获取URL的响应头信息
    headers = get_response_header(url)

    # 如果响应头信息存在
    if headers is not None:
        # 从URL中提取域名
        domain = urlparse(url).netloc
        code, title, server, location = headers

        # 打印URL和相关信息到控制台
        print(
            f"URL: {url}  Domain: {domain}    Code: {code}    Title: {title}  Server: {server}    Location: {location}")

        # 如果提供了输出文件路径
        if output_file:
            with lock:  # 使用锁确保线程安全
                try:
                    # 尝试加载已存在的工作簿
                    wb = openpyxl.load_workbook(output_file)

                except FileNotFoundError:
                    # 如果文件不存在，创建新的工作簿并添加表头
                    wb = Workbook()
                    ws = wb.active
                    headers = ["URL", "域名", "响应码", "响应标题", "网站服务器", "网站重定向"]
                    ws.append(headers)
                else:
                    # 如果文件存在，获取活动工作表
                    ws = wb.active

                # 将URL及相关信息添加到工作表中
                ws.append([url, domain, code, title, server, location])

                # 保存工作簿到指定的输出文件
                wb.save(output_file)
    else:
        # 处理异常情况，将异常的URL添加到表格中
        with lock:
            try:
                # 尝试加载已存在的工作簿
                wb = openpyxl.load_workbook(output_file)
            except FileNotFoundError:
                # 如果文件不存在，创建新的工作簿并添加表头
                wb = Workbook()
                ws = wb.active
                headers = ["URL", "域名", "响应码", "响应标题", "网站服务器", "网站重定向"]
                ws.append(headers)
            else:
                # 如果文件存在，获取活动工作表
                ws = wb.active

            # 将异常的URL信息添加到工作表中
            ws.append([url, None, None, None, None, None])

            # 保存工作簿到指定的输出文件
            wb.save(output_file)


def process_urls_from_file(file_path, output_file):
    """
    从文件中读取URL列表，对每个URL启动一个线程进行处理。
    参数:
    - file_path (str): 包含URL列表的文件路径。
    - output_file (str): 保存输出信息的Excel文件路径。
    """
    # 从文件中读取URL列表
    with open(file_path, 'r') as file:
        urls = file.readlines()
    threads = []
    # 遍历URL列表，启动线程处理每个URL
    for url in urls:
        url = url.strip()

        # 检查URL是否已经被访问过，避免重复处理
        if url.lower() not in thread_local_data.data.visited_urls:
            thread_local_data.data.visited_urls.add(url.lower())

            # 启动线程处理URL
            thread = threading.Thread(target=process_url, args=(url, output_file))
            thread.start()
            threads.append(thread)

    # 等待所有线程完成
    for thread in threads:
        thread.join()

def handle_command_line_arguments():
    # 创建参数解析器
    parser = argparse.ArgumentParser(description='存活验证')

    # 添加命令行参数
    parser.add_argument('-f', '--file', required=True, help='包含URL的输入文件。')
    parser.add_argument('-o', '--output', help='输出Excel文件。')
    # 解析命令行参数
    return parser.parse_args()

def main():
    args = handle_command_line_arguments()

    if args.output:
        if os.path.exists(args.output):
            raise FileExistsError(f"错误：输出文件 '{args.output}' 已经存在。请提供一个新的输出文件路径。")

    # 处理URLs
    process_urls_from_file(args.file, args.output)

if __name__ == "__main__":
    main()
